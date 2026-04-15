import os
import re
from copy import copy

import pandas as pd
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook


def convert_to_number(value, label=None):
    if pd.isna(value):
        return None

    raw_text = str(value).strip()
    text = raw_text.replace(",", "").strip().upper()
    label_text = "" if label is None else str(label).upper()

    if isinstance(value, (int, float)) and "%" in label_text:
        return f"{value * 100:.2f}%"

    if "%" in text:
        try:
            number = float(text.replace("%", ""))
            return f"{number:.2f}%"
        except Exception:
            return raw_text

    match = re.match(r"^(\d+\.?\d*)([KMB]?)$", text)
    if not match:
        return raw_text

    number, suffix = match.groups()
    number = float(number)

    if suffix == "K":
        number *= 1_000
    elif suffix == "M":
        number *= 1_000_000
    elif suffix == "B":
        number *= 1_000_000_000

    return round(number, 2)


def clean_sheet_name(file_path):
    name = os.path.splitext(os.path.basename(file_path))[0]
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, "_")
    return name[:31]


def copy_conditional_formatting(template_ws, ws):
    if template_ws.conditional_formatting:
        for cf_range in template_ws.conditional_formatting:
            rules = template_ws.conditional_formatting[cf_range]
            for rule in rules:
                ws.conditional_formatting.add(cf_range, copy(rule))


def normalize_text_value(value):
    if pd.isna(value):
        return ""

    text = str(value).strip()

    if text.lower() in {"nan", "none", "null"}:
        return ""

    return text


def clean_text_series(series):
    return series.apply(normalize_text_value)


def extract_company_value(source_match_col, source_return_col):
    company_value = None
    for i, label in enumerate(source_match_col):
        if str(label).strip().upper() in ["TARGET COMPANY", "COMPANY", "COMPANY NAME"]:
            company_value = source_return_col.iloc[i]
            break
    return company_value


def build_source_data(source_match_col, source_return_col):
    source_data = pd.DataFrame({
        "match_text": clean_text_series(source_match_col),
        "return_value": source_return_col,
    })

    source_data = source_data[source_data["match_text"] != ""].reset_index(drop=True)

    source_data["return_value"] = source_data.apply(
        lambda row: convert_to_number(row["return_value"], row["match_text"]),
        axis=1
    )

    return source_data


def get_valid_target_rows(target_match_col):
    target_df = pd.DataFrame({
        "original_index": target_match_col.index,
        "match_text": clean_text_series(target_match_col)
    })

    valid_target_df = target_df[target_df["match_text"] != ""].reset_index(drop=True)
    return target_df, valid_target_df


def compute_matches(source_data, target_match_col, model, threshold):
    target_df, valid_target_df = get_valid_target_rows(target_match_col)

    matched_return_values = [None] * len(target_match_col)
    matched_scores = [None] * len(target_match_col)

    if source_data.empty:
        print("WARNING: source_data is empty. No valid source match_text values found.")
        return matched_return_values, matched_scores

    if valid_target_df.empty:
        print("WARNING: target_match_col is empty. No valid target match_text values found.")
        return matched_return_values, matched_scores

    source_texts = source_data["match_text"].tolist()
    target_texts = valid_target_df["match_text"].tolist()

    print(f"Source text count: {len(source_texts)}")
    print(f"Target text count: {len(target_texts)}")

    source_embeddings = model.encode(
        source_texts,
        convert_to_tensor=True,
        normalize_embeddings=True
    )

    target_embeddings = model.encode(
        target_texts,
        convert_to_tensor=True,
        normalize_embeddings=True
    )

    similarity_matrix = util.cos_sim(target_embeddings, source_embeddings).cpu().numpy()
    best_match_idx = similarity_matrix.argmax(axis=1)
    best_match_score = similarity_matrix.max(axis=1)

    for i, score in enumerate(best_match_score):
        original_target_index = valid_target_df.iloc[i]["original_index"]

        if score >= threshold:
            idx = best_match_idx[i]
            matched_return_values[original_target_index] = source_data.iloc[idx]["return_value"]
            matched_scores[original_target_index] = float(score)
        else:
            matched_return_values[original_target_index] = None
            matched_scores[original_target_index] = float(score)

    return matched_return_values, matched_scores


def get_template_sheet(wb, preferred_name=None):
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]

    if preferred_name:
        wanted = preferred_name.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == wanted:
                return wb[name]

    if wb.worksheets:
        return wb.worksheets[0]

    raise ValueError("Target workbook has no worksheets.")


def read_excel_grid(path, sheet_name):
    return pd.read_excel(
        path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=None
    )


def get_target_match_column_from_ws(ws, start_row, end_row):
    values = []
    for row in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row, column=4).value  # Column D
        values.append(normalize_text_value(cell_value))
    return pd.Series(values)


def run_im_matching(
    input_file_source_im,
    input_file_target,
    output_path,
    input_tab_source_im="IM Summary",
    input_tab_target="Company KPI Manger",
    threshold=0.60,
    model_name="all-MiniLM-L6-v2",
):
    os.makedirs(output_path, exist_ok=True)
    model = SentenceTransformer(model_name)

    wb = load_workbook(input_file_target, keep_vba=True)
    template_ws = get_template_sheet(wb, input_tab_target)

    for source_file_im in input_file_source_im:
        print(f"Processing IM file: {source_file_im}")

        df_source_im = read_excel_grid(source_file_im, input_tab_source_im)

        source_match_col_im = clean_text_series(df_source_im.iloc[:, 1])
        source_return_col_im = df_source_im.iloc[:, 2]

        company_value = extract_company_value(source_match_col_im, source_return_col_im)

        start_row = 14
        end_row = 49

        target_match_col_im = get_target_match_column_from_ws(template_ws, start_row, end_row)
        source_data_im = build_source_data(source_match_col_im, source_return_col_im)

        print(f"IM source rows: {len(source_match_col_im)}")
        print(f"IM usable source rows: {len(source_data_im)}")
        print("IM source_data sample:")
        print(source_data_im.head(10))

        matched_return_values_im, matched_scores_im = compute_matches(
            source_data_im, target_match_col_im, model, threshold
        )

        ws = wb.copy_worksheet(template_ws)
        copy_conditional_formatting(template_ws, ws)
        ws.title = clean_sheet_name(source_file_im)

        if company_value is not None:
            ws.cell(row=3, column=2).value = company_value

        ws.cell(row=13, column=8).value = "Matched Value"
        ws.cell(row=13, column=9).value = "Similarity Score"

        for i in range(min(len(matched_return_values_im), end_row - start_row + 1)):
            excel_row = start_row + i
            ws.cell(row=excel_row, column=8).value = matched_return_values_im[i]
            ws.cell(row=excel_row, column=9).value = matched_scores_im[i]

    output_file = os.path.join(output_path, "Combined_IM_Output.xlsm")
    wb.save(output_file)
    return output_file


def run_ip_matching(
    input_file_source_ip,
    input_file_target,
    output_path,
    input_tab_source_ip="Deal KPIs (Finance Use)",
    input_tab_target="Company KPI Manger",
    threshold=0.60,
    model_name="all-MiniLM-L6-v2",
):
    os.makedirs(output_path, exist_ok=True)
    model = SentenceTransformer(model_name)

    wb = load_workbook(input_file_target, keep_vba=True)
    template_ws = get_template_sheet(wb, input_tab_target)

    for source_file_ip in input_file_source_ip:
        print(f"Processing IP file: {source_file_ip}")

        df_source_ip = read_excel_grid(source_file_ip, input_tab_source_ip)

        source_match_col_ip = clean_text_series(df_source_ip.iloc[:, 1])
        source_match_col_ip = source_match_col_ip.str.replace(r"\bCARR\b", "ARR", regex=True)
        source_return_col_ip = df_source_ip.iloc[:, 2]

        company_value_ip = extract_company_value(source_match_col_ip, source_return_col_ip)

        start_row = 55
        end_row = 106

        target_match_col_ip = get_target_match_column_from_ws(template_ws, start_row, end_row)
        source_data_ip = build_source_data(source_match_col_ip, source_return_col_ip)

        print(f"IP source rows: {len(source_match_col_ip)}")
        print(f"IP usable source rows: {len(source_data_ip)}")
        print("IP source_data sample:")
        print(source_data_ip.head(10))

        matched_return_values_ip, matched_scores_ip = compute_matches(
            source_data_ip, target_match_col_ip, model, threshold
        )

        ws = wb.copy_worksheet(template_ws)
        copy_conditional_formatting(template_ws, ws)
        ws.title = clean_sheet_name(source_file_ip)

        if company_value_ip is not None:
            ws.cell(row=3, column=2).value = company_value_ip

        ws.cell(row=54, column=8).value = "Matched Value"
        ws.cell(row=54, column=9).value = "Similarity Score"

        for i in range(min(len(matched_return_values_ip), end_row - start_row + 1)):
            excel_row = start_row + i
            ws.cell(row=excel_row, column=8).value = matched_return_values_ip[i]
            ws.cell(row=excel_row, column=9).value = matched_scores_ip[i]

    output_file = os.path.join(output_path, "Combined_IP_Output.xlsm")
    wb.save(output_file)
    return output_file
